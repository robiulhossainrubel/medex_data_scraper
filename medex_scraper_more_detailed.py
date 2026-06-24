import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import os
import re

OUTPUT_FILE = "medex_medicines_more_detailed.xlsx"
BASE_URL = "https://medex.com.bd/brands"
BATCH_SIZE = 5
LIST_DELAY = 1.0
DETAIL_DELAY = 1.5

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

# All possible section IDs found across different medicine pages
SECTION_MAP = {
    "indications":           "Indications",
    "description":           "Description",          # topical/misc medicines
    "composition":           "Composition",
    "compound_summary":      "Compound Summary",     # some generics
    "mode_of_action":        "Pharmacology",
    "dosage":                "Dosage & Administration",
    "administration":        "Administration",        # alternate dosage section
    "reconstitution":        "Reconstitution",        # injectable medicines
    "pediatric_uses":        "Pediatric Uses",        # some medicines
    "interaction":           "Interaction",
    "contraindications":     "Contraindications",
    "side_effects":          "Side Effects",
    "pregnancy_cat":         "Pregnancy & Lactation",
    "precautions":           "Precautions & Warnings",
    "overdose_effects":      "Overdose Effects",
    "drug_classes":          "Therapeutic Class",
    "storage_conditions":    "Storage Conditions",
}

COLUMNS = [
    "Name", "Type", "Dose", "Generic", "Company",
    "Price Label", "Price", "Strip Price", "Pack Size",
    "Indications", "Description", "Composition", "Compound Summary",
    "Pharmacology", "Dosage & Administration", "Administration",
    "Reconstitution", "Pediatric Uses",
    "Interaction", "Contraindications", "Side Effects",
    "Pregnancy & Lactation", "Precautions & Warnings",
    "Overdose Effects", "Therapeutic Class", "Storage Conditions",
    "URL"
]


def setup_workbook():
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        next_row = ws.max_row + 1
        total = next_row - 2
        print(f"Resuming: {total} records already saved.")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Medicines"

        header_fill = PatternFill("solid", start_color="1B5E20")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, h in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Column widths
        widths = {
            "Name": 22, "Type": 18, "Dose": 20, "Generic": 30, "Company": 28,
            "Price Label": 18, "Price": 12, "Strip Price": 12, "Pack Size": 20,
            "Indications": 50, "Description": 50, "Composition": 50,
            "Compound Summary": 40, "Pharmacology": 50,
            "Dosage & Administration": 50, "Administration": 40,
            "Reconstitution": 35, "Pediatric Uses": 35,
            "Interaction": 40, "Contraindications": 40, "Side Effects": 40,
            "Pregnancy & Lactation": 35, "Precautions & Warnings": 40,
            "Overdose Effects": 35, "Therapeutic Class": 25,
            "Storage Conditions": 30, "URL": 60,
        }
        for col, h in enumerate(COLUMNS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = widths.get(h, 25)

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        next_row = 2

    return wb, ws, next_row


def get_total_pages():
    r = fetch_with_retry(BASE_URL)
    if not r:
        return 1
    soup = BeautifulSoup(r.text, "html.parser")
    nums = []
    for a in soup.select("ul.pagination li.page-item a.page-link"):
        try:
            nums.append(int(a.text.strip()))
        except ValueError:
            pass
    return max(nums) if nums else 1


def get_page_urls(page_num):
    """Collect all medicine URLs from one list page."""
    r = fetch_with_retry(f"{BASE_URL}?page={page_num}")
    if not r:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    return [a.get("href", "").strip() for a in soup.select("a.hoverable-block") if a.get("href")]


def clean(text):
    """Normalize whitespace."""
    return re.sub(r'\s+', ' ', text).strip() if text else ""


def parse_price(soup):
    """
    Handle all price label variations:
    - "Unit Price: ৳ 12.00" with optional strip price & pack size
    - "0.5 ml ampoule: ৳ 25.00" with optional pack size
    - "100 ml bottle: ৳ 195.00" etc.
    Returns: price_label, price, strip_price, pack_size
    """
    price_label = ""
    price = ""
    strip_price = ""
    pack_size = ""

    pkg = soup.select_one(".package-container")
    if not pkg:
        return price_label, price, strip_price, pack_size

    # Price label (the colored span — e.g. "Unit Price:", "100 ml bottle:", etc.)
    label_span = pkg.select_one("span[style]")
    if label_span:
        price_label = clean(label_span.get_text()).rstrip(":")

    # Price value (the plain ৳ span right after label)
    spans = pkg.find_all("span", recursive=False)
    for sp in spans:
        if not sp.get("style") and "৳" in sp.get_text():
            price = clean(sp.get_text())
            break

    # Strip price — "Strip Price: ৳ 120.00"
    pkg_text = pkg.get_text(separator=" ")
    sp_match = re.search(r'Strip Price[:\s]*৳\s*([\d,\.]+)', pkg_text)
    if sp_match:
        strip_price = "৳ " + sp_match.group(1)

    # Pack size — e.g. "(6 x 10: ৳ 720.00)" or "(12's pack: ৳ 420.00)"
    ps_match = re.search(r'\(([^)]+)\)', pkg_text)
    if ps_match:
        pack_size = clean(ps_match.group(1))

    return price_label, price, strip_price, pack_size


def parse_detail_page(soup, url):
    data = {col: "" for col in COLUMNS}
    data["URL"] = url

    # --- Name & Type ---
    h1 = soup.select_one("h1.page-heading-1-l.brand")
    if h1:
        small = h1.select_one("small")
        data["Type"] = clean(small.get_text()) if small else ""
        if small:
            small.decompose()
        data["Name"] = clean(h1.get_text())

    # --- Generic ---
    generic_div = soup.find("div", title="Generic Name")
    if generic_div:
        data["Generic"] = clean(generic_div.get_text())

    # --- Dose / Strength ---
    strength_div = soup.find("div", title="Strength")
    if strength_div:
        data["Dose"] = clean(strength_div.get_text())

    # --- Company ---
    # The first "Manufactured by" div has the company link
    for div in soup.find_all("div", title="Manufactured by"):
        a = div.select_one("a.calm-link")
        if a:
            data["Company"] = clean(a.get_text())
            break

    # --- Price (all variations) ---
    data["Price Label"], data["Price"], data["Strip Price"], data["Pack Size"] = parse_price(soup)

    # --- All content sections ---
    for section_id, col_name in SECTION_MAP.items():
        anchor = soup.find(id=section_id)
        if anchor:
            # ac-body is always the next sibling div of the anchor div
            body = anchor.find_next_sibling("div", class_="ac-body")
            if body:
                data[col_name] = clean(body.get_text(separator=" "))

    return data


def fetch_with_retry(url, retries=3, backoff=5):
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=20)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            if attempt < retries - 1:
                print(f"\n  Retry {attempt+1} for {url} ({e})")
                time.sleep(backoff * (attempt + 1))
            else:
                print(f"\n  FAILED: {url} — {e}")
    return None


def write_batch(ws, records, start_row):
    even_fill = PatternFill("solid", start_color="F1F8E9")
    odd_fill  = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    content_cols = set(range(10, len(COLUMNS) + 1))  # cols 10 onwards = content

    for i, rec in enumerate(records):
        row = start_row + i
        fill = even_fill if i % 2 == 0 else odd_fill
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = rec.get(col_name, "")
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            wrap = col_idx in content_cols
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
        ws.row_dimensions[row].height = 55


def main():
    print("=" * 60)
    print("MedEx Scraper — Page by Page | Full Detail Data")
    print("=" * 60)

    wb, ws, next_row = setup_workbook()
    total_scraped = next_row - 2
    batch_buffer = []

    print("Fetching total pages...")
    total_pages = get_total_pages()
    print(f"Total pages: {total_pages}\n")

    global_count = 0  # tracks total medicines seen so far across pages

    for page in range(1, total_pages + 1):

        # Step 1: Collect URLs from this page
        print(f"{'='*55}")
        print(f"PAGE {page}/{total_pages} — Collecting URLs...")
        urls = get_page_urls(page)
        count_on_page = len(urls)
        print(f"  {count_on_page} medicines found on page {page}")

        if not urls:
            print("  No URLs, skipping page.")
            time.sleep(LIST_DELAY)
            continue

        page_end = global_count + count_on_page

        # Skip fully scraped pages
        if page_end <= total_scraped:
            print(f"  Already scraped (#{global_count+1}–{page_end}), skipping.")
            global_count += count_on_page
            time.sleep(LIST_DELAY)
            continue

        # Partial resume within a page
        skip_in_page = max(0, total_scraped - global_count)
        if skip_in_page > 0:
            print(f"  Skipping first {skip_in_page} (already done).")
        urls_to_do = urls[skip_in_page:]

        time.sleep(LIST_DELAY)

        # Step 2: Process each detail page on this page
        for i, url in enumerate(urls_to_do, 1):
            abs_num = global_count + skip_in_page + i
            name_hint = url.split("/")[-1]
            print(f"  [{abs_num}/{total_pages*count_on_page}] {name_hint}", end="  ")

            r = fetch_with_retry(url)
            if r:
                soup = BeautifulSoup(r.text, "html.parser")
                record = parse_detail_page(soup, url)
                batch_buffer.append(record)
                print(f"✓ {record.get('Name','?')} | {record.get('Type','?')} | {record.get('Price','')}")
            else:
                empty = {c: "" for c in COLUMNS}
                empty["URL"] = url
                batch_buffer.append(empty)
                print("✗ (failed, saved as empty row)")

            time.sleep(DETAIL_DELAY)

            # Save every BATCH_SIZE records
            if len(batch_buffer) >= BATCH_SIZE:
                print(f"\n  >>> Saving {BATCH_SIZE} records...")
                write_batch(ws, batch_buffer[:BATCH_SIZE], next_row)
                next_row += BATCH_SIZE
                total_scraped += BATCH_SIZE
                batch_buffer = batch_buffer[BATCH_SIZE:]
                wb.save(OUTPUT_FILE)
                print(f"  >>> Saved! Total in file: {total_scraped}\n")

        global_count += count_on_page
        print(f"  Page {page} complete. Total scraped: {total_scraped + len(batch_buffer)}")

    # Save remaining
    if batch_buffer:
        print(f"\nSaving final {len(batch_buffer)} records...")
        write_batch(ws, batch_buffer, next_row)
        total_scraped += len(batch_buffer)
        wb.save(OUTPUT_FILE)

    print(f"\n{'='*60}")
    print(f"DONE! Total: {total_scraped} medicines → {OUTPUT_FILE}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()