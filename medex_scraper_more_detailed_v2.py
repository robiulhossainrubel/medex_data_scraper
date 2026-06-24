import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import os
import re
import random
import glob

OUTPUT_FILE = "medex_medicines_more_detailed_v2.xlsx"
TEMP_FILE_PATTERN = OUTPUT_FILE.replace(".xlsx", "_temp*.xlsx")   # pattern for glob
PROGRESS_FILE = "scraper_progress.txt"
BASE_URL = "https://medex.com.bd/brands"
BATCH_SIZE = 5
LIST_DELAY = 1.0
DETAIL_DELAY = 1.5

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
]

BASE_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

SECTION_MAP = {
    "indications": "Indications",
    "description": "Description",
    "composition": "Composition",
    "compound_summary": "Compound Summary",
    "mode_of_action": "Pharmacology",
    "dosage": "Dosage & Administration",
    "administration": "Administration",
    "reconstitution": "Reconstitution",
    "pediatric_uses": "Pediatric Uses",
    "interaction": "Interaction",
    "contraindications": "Contraindications",
    "side_effects": "Side Effects",
    "pregnancy_cat": "Pregnancy & Lactation",
    "precautions": "Precautions & Warnings",
    "overdose_effects": "Overdose Effects",
    "drug_classes": "Therapeutic Class",
    "storage_conditions": "Storage Conditions",
}

COLUMNS = [
    "Name", "Type", "Dose", "Generic", "Company",
    "Price Label", "Price", "Strip Price", "Pack Size",
    "Indications", "Description", "Composition", "Compound Summary",
    "Pharmacology", "Dosage & Administration", "Administration",
    "Reconstitution", "Pediatric Uses",
    "Interaction", "Contraindications", "Side Effects",
    "Pregnancy & Lactation", "Precautions & Warnings",
    "Overdose Effects", "Therapeutic Class", "Storage Conditions",
    "URL"
]


# ---------- workbook & progress helpers ----------
def setup_workbook():
    """Excel ওয়ার্কবুক তৈরি বা লোড করে। ফাইল খোলা থাকলে পুনরায় চেষ্টা করে।"""
    for attempt in range(5):
        try:
            if os.path.exists(OUTPUT_FILE):
                wb = openpyxl.load_workbook(OUTPUT_FILE)
                ws = wb.active
                next_row = ws.max_row + 1
                print(f"Resuming: {next_row - 2} records already in Excel.")
                return wb, ws, next_row
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Medicines"
                header_fill = PatternFill("solid", start_color="1B5E20")
                header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for col, h in enumerate(COLUMNS, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                widths = {
                    "Name": 22, "Type": 18, "Dose": 20, "Generic": 30, "Company": 28,
                    "Price Label": 18, "Price": 12, "Strip Price": 12, "Pack Size": 20,
                    "Indications": 50, "Description": 50, "Composition": 50,
                    "Compound Summary": 40, "Pharmacology": 50,
                    "Dosage & Administration": 50, "Administration": 40,
                    "Reconstitution": 35, "Pediatric Uses": 35,
                    "Interaction": 40, "Contraindications": 40, "Side Effects": 40,
                    "Pregnancy & Lactation": 35, "Precautions & Warnings": 40,
                    "Overdose Effects": 35, "Therapeutic Class": 25,
                    "Storage Conditions": 30, "URL": 60,
                }
                for col, h in enumerate(COLUMNS, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = widths.get(h, 25)
                ws.row_dimensions[1].height = 22
                ws.freeze_panes = "A2"
                return wb, ws, 2
        except PermissionError:
            print(f"'{OUTPUT_FILE}' খোলা আছে। বন্ধ করে দিন, ৫ সেকেন্ড অপেক্ষা করছি...")
            time.sleep(5)
    raise PermissionError(f"'{OUTPUT_FILE}' বন্ধ করা সম্ভব হয়নি। স্ক্রিপ্ট বন্ধ হচ্ছে।")


def load_progress():
    """প্রগ্রেস ফাইল থেকে (পেজ, করা_আইটেম, মোট_স্ক্র্যাপ) পড়ে।"""
    if not os.path.exists(PROGRESS_FILE):
        return 1, 0, 0
    try:
        with open(PROGRESS_FILE, "r") as f:
            parts = f.read().strip().split(",")
            if len(parts) == 3:
                page, done, total = map(int, parts)
                return page, done, total
            elif len(parts) == 2:   # পুরনো ফরম্যাট (compatibility)
                page, done = map(int, parts)
                return page, done, 0
    except Exception:
        pass
    return 1, 0, 0


def save_progress(page, items_done, total_scraped):
    """প্রগ্রেস ফাইলে বর্তমান অবস্থা লেখে।"""
    with open(PROGRESS_FILE, "w") as f:
        f.write(f"{page},{items_done},{total_scraped}")


# ---------- temp file sync ----------
def sync_all_temp_files(wb, ws, expected_records):
    """
    সমস্ত '_temp*.xlsx' ফাইল চেক করে অনুপস্থিত রেকর্ড মূল শিটে জুড়ে দেয়।
    তারপর সব টেম্প ফাইল মুছে ফেলে।
    """
    temp_files = sorted(glob.glob(OUTPUT_FILE.replace(".xlsx", "_temp*.xlsx")))
    if not temp_files:
        return

    actual_records = ws.max_row - 1  # header বাদে
    if actual_records >= expected_records:
        # সব রেকর্ডই আছে, অপ্রয়োজনীয় টেম্প ফাইলগুলো মুছে দিন
        for tf in temp_files:
            try:
                os.remove(tf)
                print(f"অপ্রয়োজনীয় টেম্প ফাইল মুছে ফেলা হয়েছে: {tf}")
            except:
                pass
        return

    missing = expected_records - actual_records
    print(f"Auto-sync: {missing} টি রেকর্ড টেম্প থেকে মূল ফাইলে যোগ করা হচ্ছে...")

    # শেষ টেম্প ফাইলটি থেকে প্রথমে নেওয়ার চেষ্টা করি (সবচেয়ে সাম্প্রতিক)
    for tf in reversed(temp_files):
        try:
            twb = openpyxl.load_workbook(tf)
            tws = twb.active
            last_row = tws.max_row
            # কত রেকর্ড নেব এই ফাইল থেকে
            records_in_file = min(missing, last_row - 1)  # header বাদ
            start = last_row - records_in_file + 1
            for r in range(start, last_row + 1):
                row_vals = [tws.cell(row=r, column=c).value for c in range(1, len(COLUMNS) + 1)]
                ws.append(row_vals)
            twb.close()
            missing -= records_in_file
            if missing <= 0:
                break
        except Exception as e:
            print(f"'{tf}' প্রসেস করতে সমস্যা: {e}")

    # মূল ফাইল সেভ
    try:
        wb.save(OUTPUT_FILE)
    except PermissionError:
        # মূল ফাইল এখনও লক? তাহলে টেম্পেই থাকবে, পরে আবার চেষ্টা হবে
        print("সিঙ্ক করার সময় মূল ফাইল লক ছিল। পরবর্তী রানে আবার চেষ্টা হবে।")
        return

    # সব টেম্প ফাইল মুছে দিন
    for tf in temp_files:
        try:
            os.remove(tf)
        except:
            pass
    print("সকল টেম্প ফাইল থেকে ডেটা মূল ফাইলে একীভূত ও পরিষ্কার করা হয়েছে।")


# ---------- scraping functions ----------
def get_total_pages():
    r = fetch_with_retry(BASE_URL)
    if not r:
        return 1
    soup = BeautifulSoup(r.text, "html.parser")
    nums = []
    for a in soup.select("ul.pagination li.page-item a.page-link"):
        try:
            nums.append(int(a.text.strip()))
        except ValueError:
            pass
    return max(nums) if nums else 1


def get_page_urls(page_num):
    r = fetch_with_retry(f"{BASE_URL}?page={page_num}")
    if not r:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    return [a.get("href", "").strip() for a in soup.select("a.hoverable-block") if a.get("href")]


def clean(text):
    return re.sub(r'\s+', ' ', text).strip() if text else ""


def parse_price(soup):
    price_label = ""
    price = ""
    strip_price = ""
    pack_size = ""
    pkg = soup.select_one(".package-container")
    if not pkg:
        return price_label, price, strip_price, pack_size
    label_span = pkg.select_one("span[style]")
    if label_span:
        price_label = clean(label_span.get_text()).rstrip(":")
    spans = pkg.find_all("span", recursive=False)
    for sp in spans:
        if not sp.get("style") and "৳" in sp.get_text():
            price = clean(sp.get_text())
            break
    pkg_text = pkg.get_text(separator=" ")
    sp_match = re.search(r'Strip Price[:\s]*৳\s*([\d,\.]+)', pkg_text)
    if sp_match:
        strip_price = "৳ " + sp_match.group(1)
    ps_match = re.search(r'\(([^)]+)\)', pkg_text)
    if ps_match:
        pack_size = clean(ps_match.group(1))
    return price_label, price, strip_price, pack_size


def parse_detail_page(soup, url):
    data = {col: "" for col in COLUMNS}
    data["URL"] = url
    h1 = soup.select_one("h1.page-heading-1-l.brand")
    if h1:
        small = h1.select_one("small")
        data["Type"] = clean(small.get_text()) if small else ""
        if small:
            small.decompose()
        data["Name"] = clean(h1.get_text())
    generic_div = soup.find("div", title="Generic Name")
    if generic_div:
        data["Generic"] = clean(generic_div.get_text())
    strength_div = soup.find("div", title="Strength")
    if strength_div:
        data["Dose"] = clean(strength_div.get_text())
    for div in soup.find_all("div", title="Manufactured by"):
        a = div.select_one("a.calm-link")
        if a:
            data["Company"] = clean(a.get_text())
            break
    data["Price Label"], data["Price"], data["Strip Price"], data["Pack Size"] = parse_price(soup)
    for section_id, col_name in SECTION_MAP.items():
        anchor = soup.find(id=section_id)
        if anchor:
            body = anchor.find_next_sibling("div", class_="ac-body")
            if body:
                data[col_name] = clean(body.get_text(separator=" "))
    return data


def fetch_with_retry(url, retries=3, backoff=5):
    for attempt in range(retries):
        headers = BASE_HEADERS.copy()
        headers["User-Agent"] = random.choice(USER_AGENTS)
        try:
            r = requests.get(url, headers=headers, timeout=20)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            if attempt < retries - 1:
                wait = backoff * (attempt + 1) + random.uniform(1, 3)
                print(f"\n  Retry {attempt+1} for {url} ({e}) – waiting {wait:.1f}s")
                time.sleep(wait)
            else:
                print(f"\n  FAILED: {url} — {e}")
    return None


def write_batch(ws, records, start_row):
    even_fill = PatternFill("solid", start_color="F1F8E9")
    odd_fill  = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    content_cols = set(range(10, len(COLUMNS) + 1))
    for i, rec in enumerate(records):
        row = start_row + i
        fill = even_fill if i % 2 == 0 else odd_fill
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = rec.get(col_name, "")
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            wrap = col_idx in content_cols
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
        ws.row_dimensions[row].height = 55


# ---------- safe save with fallback temp ----------
def safe_save(wb, filename, max_retries=5):
    """
    PermissionError পেলে প্রথমে মূল ফাইল, তারপর '_temp.xlsx', তারপর '_temp_1.xlsx' …
    এভাবে চেষ্টা করে। যে ফাইল তৈরি হয় তার নাম ফেরত দেয়। সম্পূর্ণ ব্যর্থ হলে None।
    """
    for attempt in range(max_retries):
        try:
            wb.save(filename)
            return filename
        except PermissionError:
            if attempt < max_retries - 1:
                wait = 5 * (attempt + 1)
                print(f"\n  ⚠️  '{filename}' সংরক্ষণে সমস্যা (ফাইল খোলা?)। {wait} সেকেন্ড অপেক্ষা...")
                time.sleep(wait)
            else:
                # মূল ফাইল সম্ভব নয়, টেম্প ফাইলে চেষ্টা
                base_temp = filename.replace(".xlsx", "_temp.xlsx")
                # প্রথম চেষ্টা সাধারণ _temp.xlsx, তারপর _temp_1.xlsx, _temp_2.xlsx ...
                for temp_idx in range(10):
                    if temp_idx == 0:
                        temp_name = base_temp
                    else:
                        temp_name = base_temp.replace(".xlsx", f"_{temp_idx}.xlsx")
                    try:
                        wb.save(temp_name)
                        print(f"\n  ❌ মূল ফাইল লক। '{temp_name}' নামে অস্থায়ী ফাইল সংরক্ষিত হয়েছে।")
                        return temp_name
                    except PermissionError:
                        if temp_idx < 9:
                            print(f"     '{temp_name}' ও খোলা। পরবর্তী নাম চেষ্টা করা হচ্ছে...")
                            time.sleep(2)
                        else:
                            print("  ❌❌ সব ফাইল লক! ডেটা সংরক্ষণ একেবারেই সম্ভব হয়নি।")
                            return None
    return None


# ---------- main ----------
def main():
    print("=" * 60)
    print("MedEx Scraper — Auto Sync + Safe Save + Random Delays")
    print("=" * 60)

    wb, ws, next_row = setup_workbook()
    start_page, skip_on_page, prev_total = load_progress()
    print(f"Progress: page {start_page}, done {skip_on_page}, total scraped earlier = {prev_total}\n")

    # শুরুতেই সমস্ত টেম্প ফাইল থেকে অনুপস্থিত ডেটা মূল ফাইলে একীভূত করুন
    sync_all_temp_files(wb, ws, prev_total)
    # সিঙ্কের পর next_row আপডেট (সারি সংখ্যা বেড়ে থাকতে পারে)
    next_row = ws.max_row + 1
    # প্রগ্রেসে সংরক্ষিত মোট রেকর্ড আর শিটের বাস্তব রেকর্ড সিঙ্ক করুন
    actual_now = next_row - 2
    if actual_now > prev_total:
        prev_total = actual_now

    total_pages = get_total_pages()
    print(f"Total pages on site: {total_pages}\n")

    if start_page > total_pages:
        print("Progress out of range, resetting.")
        start_page = 1
        skip_on_page = 0
        prev_total = 0
        save_progress(1, 0, 0)

    batch_buffer = []
    total_scraped = prev_total

    for page in range(start_page, total_pages + 1):
        print(f"{'='*55}")
        print(f"PAGE {page}/{total_pages} — collecting URLs...")
        urls = get_page_urls(page)
        count_on_page = len(urls)
        print(f"  {count_on_page} medicines found on page {page}")

        if not urls:
            time.sleep(LIST_DELAY + random.uniform(0.5, 2.0))
            save_progress(page + 1, 0, total_scraped)
            continue

        if page == start_page:
            start_offset = skip_on_page
            if start_offset >= len(urls):
                print("  This page already fully done, moving on.")
                save_progress(page + 1, 0, total_scraped)
                time.sleep(LIST_DELAY + random.uniform(0.5, 2.0))
                continue
            urls_to_do = urls[start_offset:]
            items_done_on_page = start_offset
            print(f"  Skipping first {start_offset} (already done).")
        else:
            start_offset = 0
            urls_to_do = urls
            items_done_on_page = 0

        time.sleep(LIST_DELAY + random.uniform(0.5, 2.0))

        for i, url in enumerate(urls_to_do, 1):
            item_num = start_offset + i
            name_hint = url.split("/")[-1]
            print(f"  [{item_num}/{count_on_page}] {name_hint}", end="  ")

            r = fetch_with_retry(url)
            if r:
                soup = BeautifulSoup(r.text, "html.parser")
                record = parse_detail_page(soup, url)
                batch_buffer.append(record)
                print(f"✓ {record.get('Name','?')} | {record.get('Type','?')} | {record.get('Price','')}")
            else:
                empty = {c: "" for c in COLUMNS}
                empty["URL"] = url
                batch_buffer.append(empty)
                print("✗ (failed)")

            time.sleep(DETAIL_DELAY + random.uniform(1.0, 3.0))

            if len(batch_buffer) >= BATCH_SIZE:
                print(f"\n  >>> Saving {BATCH_SIZE} records...")
                write_batch(ws, batch_buffer[:BATCH_SIZE], next_row)
                next_row += BATCH_SIZE
                batch_buffer = batch_buffer[BATCH_SIZE:]

                saved_to = safe_save(wb, OUTPUT_FILE)
                if saved_to == OUTPUT_FILE:
                    # মূল ফাইলে সফল সেভ, টেম্প ফাইল থাকলে পরিষ্কার করুন
                    for tf in glob.glob(TEMP_FILE_PATTERN):
                        try:
                            os.remove(tf)
                        except:
                            pass
                    # print("টেম্প ফাইল ক্লিনআপ সম্পন্ন")  # optional
                elif saved_to:
                    # টেম্প ফাইলে সেভ হয়েছে, পরের রানে সিঙ্ক হবে
                    pass
                else:
                    # একেবারে ব্যর্থ, তবুও প্রগ্রেস সেভ করি
                    print("  ⚠️ সেভ ব্যর্থ, প্রগ্রেস সংরক্ষণ করা হচ্ছে...")

                items_done_on_page += BATCH_SIZE
                total_scraped += BATCH_SIZE
                save_progress(page, items_done_on_page, total_scraped)
                print(f"  >>> Saved! Page {page}, {items_done_on_page} done, total {total_scraped}\n")

        # বাকি রেকর্ড (BATCH_SIZE এর চেয়ে কম) থাকলে সেভ করুন
        if batch_buffer:
            remaining = len(batch_buffer)
            write_batch(ws, batch_buffer, next_row)
            next_row += remaining
            saved_to = safe_save(wb, OUTPUT_FILE)
            if saved_to == OUTPUT_FILE:
                for tf in glob.glob(TEMP_FILE_PATTERN):
                    try:
                        os.remove(tf)
                    except:
                        pass
            items_done_on_page += remaining
            total_scraped += remaining
            batch_buffer = []
            print(f"  Saved last {remaining} records of page.")

        # পেজ শেষ, পরবর্তী পেজের প্রগ্রেস লিখি
        save_progress(page + 1, 0, total_scraped)
        print(f"  Page {page} complete. Progress updated to page {page+1}.\n")

    # সম্পূর্ণ শেষে টেম্প ফাইল থাকলে পরিষ্কার
    for tf in glob.glob(TEMP_FILE_PATTERN):
        try:
            os.remove(tf)
        except:
            pass
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)
    print(f"\n{'='*60}")
    print(f"DONE! Total medicines scraped: {total_scraped} → {OUTPUT_FILE}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()