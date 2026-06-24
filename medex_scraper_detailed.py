import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import os
import re

OUTPUT_FILE = "medex_medicines_detailed.xlsx"
BASE_URL = "https://medex.com.bd/brands"
BATCH_SIZE = 5
LIST_DELAY = 1.0
DETAIL_DELAY = 1.5

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

COLUMNS = [
    "Name", "Type", "Dose", "Generic", "Company",
    "Unit Price", "Strip Price", "Pack Size",
    "Indications", "Composition", "Pharmacology",
    "Dosage & Administration", "Interaction", "Contraindications",
    "Side Effects", "Pregnancy & Lactation", "Precautions & Warnings",
    "Overdose Effects", "Therapeutic Class", "Storage Conditions",
    "URL"
]

SECTION_MAP = {
    "indications":        "Indications",
    "composition":        "Composition",
    "mode_of_action":     "Pharmacology",
    "dosage":             "Dosage & Administration",
    "interaction":        "Interaction",
    "contraindications":  "Contraindications",
    "side_effects":       "Side Effects",
    "pregnancy_cat":      "Pregnancy & Lactation",
    "precautions":        "Precautions & Warnings",
    "overdose_effects":   "Overdose Effects",
    "drug_classes":       "Therapeutic Class",
    "storage_conditions": "Storage Conditions",
}


def setup_workbook():
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        next_row = ws.max_row + 1
        total = next_row - 2
        print(f"Resuming: {total} records already saved.")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Medicines"

        header_fill = PatternFill("solid", start_color="1B5E20")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, h in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        col_widths = [22, 18, 20, 30, 30, 12, 12, 18,
                      45, 45, 45, 45, 35, 35, 30, 30, 35, 30, 25, 30, 60]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        next_row = 2

    return wb, ws, next_row


def get_total_pages():
    r = fetch_with_retry(BASE_URL)
    if not r:
        return 1
    soup = BeautifulSoup(r.text, "html.parser")
    nums = []
    for a in soup.select("ul.pagination li.page-item a.page-link"):
        try:
            nums.append(int(a.text.strip()))
        except ValueError:
            pass
    return max(nums) if nums else 1


def get_page_urls(page_num):
    """Get all medicine URLs from a single list page."""
    r = fetch_with_retry(f"{BASE_URL}?page={page_num}")
    if not r:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    urls = []
    for a in soup.select("a.hoverable-block"):
        href = a.get("href", "").strip()
        if href:
            urls.append(href)
    return urls


def clean(text):
    return re.sub(r'\s+', ' ', text).strip() if text else ""


def parse_detail_page(soup, url):
    data = {col: "" for col in COLUMNS}
    data["URL"] = url

    # Name & Type
    h1 = soup.select_one("h1.page-heading-1-l.brand")
    if h1:
        small = h1.select_one("small")
        data["Type"] = clean(small.get_text()) if small else ""
        if small:
            small.decompose()
        data["Name"] = clean(h1.get_text())

    # Generic
    generic_div = soup.find("div", title="Generic Name")
    if generic_div:
        data["Generic"] = clean(generic_div.get_text())

    # Dose / Strength
    strength_div = soup.find("div", title="Strength")
    if strength_div:
        data["Dose"] = clean(strength_div.get_text())

    # Company
    company_div = soup.find("div", title="Manufactured by")
    if company_div:
        a = company_div.select_one("a.calm-link")
        if a:
            data["Company"] = clean(a.get_text())

    # Prices
    pkg = soup.select_one(".package-container")
    if pkg:
        pkg_text = pkg.get_text(separator=" ", strip=True)
        up = re.search(r'Unit Price[:\s]*৳\s*([\d,\.]+)', pkg_text)
        if up:
            data["Unit Price"] = "৳ " + up.group(1)
        sp = re.search(r'Strip Price[:\s]*৳\s*([\d,\.]+)', pkg_text)
        if sp:
            data["Strip Price"] = "৳ " + sp.group(1)
        ps = re.search(r'\(([^)]+৳[^)]+)\)', pkg_text)
        if ps:
            data["Pack Size"] = ps.group(1)

    # Content sections
    for section_id, col_name in SECTION_MAP.items():
        anchor = soup.find(id=section_id)
        if anchor:
            body = anchor.find_next_sibling("div", class_="ac-body")
            if body:
                data[col_name] = clean(body.get_text(separator=" "))

    return data


def fetch_with_retry(url, retries=3, backoff=5):
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=20)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            if attempt < retries - 1:
                time.sleep(backoff * (attempt + 1))
            else:
                print(f"\n  FAILED: {url} — {e}")
    return None


def write_batch(ws, records, start_row):
    even_fill = PatternFill("solid", start_color="F1F8E9")
    odd_fill  = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=9)

    for i, rec in enumerate(records):
        row = start_row + i
        fill = even_fill if i % 2 == 0 else odd_fill
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = rec.get(col_name, "")
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            wrap = col_idx >= 9 or col_idx == len(COLUMNS)
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
        ws.row_dimensions[row].height = 55


def main():
    print("=" * 60)
    print("MedEx Scraper — Page by Page")
    print("=" * 60)

    wb, ws, next_row = setup_workbook()
    total_scraped = next_row - 2
    batch_buffer = []

    print("Fetching total pages...")
    total_pages = get_total_pages()
    print(f"Total pages: {total_pages}\n")

    # Figure out which page to resume from
    # Track globally how many medicines we've passed
    global_medicine_count = 0

    for page in range(1, total_pages + 1):

        # --- Step 1: Collect URLs from this page ---
        print(f"{'='*50}")
        print(f"PAGE {page}/{total_pages} — Collecting URLs...")
        urls = get_page_urls(page)
        print(f"  Found {len(urls)} medicines on page {page}")

        if not urls:
            print("  No URLs found, skipping.")
            time.sleep(LIST_DELAY)
            continue

        # Skip pages already fully scraped
        page_end = global_medicine_count + len(urls)
        if page_end <= total_scraped:
            print(f"  Already scraped (records {global_medicine_count+1}–{page_end}), skipping.")
            global_medicine_count += len(urls)
            time.sleep(LIST_DELAY)
            continue

        # Partial resume: skip already-done medicines within this page
        skip_in_page = max(0, total_scraped - global_medicine_count)
        if skip_in_page > 0:
            print(f"  Skipping first {skip_in_page} already scraped.")
        urls_to_process = urls[skip_in_page:]

        time.sleep(LIST_DELAY)

        # --- Step 2: Process each medicine detail page ---
        for i, url in enumerate(urls_to_process, 1):
            abs_num = global_medicine_count + skip_in_page + i
            print(f"  [{abs_num}] {url.split('/')[-1]}", end="  ")

            r = fetch_with_retry(url)
            if r:
                soup = BeautifulSoup(r.text, "html.parser")
                record = parse_detail_page(soup, url)
                batch_buffer.append(record)
                print(f"✓  {record.get('Name', '?')}")
            else:
                empty = {c: "" for c in COLUMNS}
                empty["URL"] = url
                batch_buffer.append(empty)
                print("✗  (failed)")

            time.sleep(DETAIL_DELAY)

            # Save every BATCH_SIZE records
            if len(batch_buffer) >= BATCH_SIZE:
                print(f"\n  >>> Saving {BATCH_SIZE} records to Excel...")
                write_batch(ws, batch_buffer[:BATCH_SIZE], next_row)
                next_row += BATCH_SIZE
                total_scraped += BATCH_SIZE
                batch_buffer = batch_buffer[BATCH_SIZE:]
                wb.save(OUTPUT_FILE)
                print(f"  >>> Saved! Total in file: {total_scraped}\n")

        global_medicine_count += len(urls)
        print(f"  Page {page} done. Total scraped so far: {total_scraped + len(batch_buffer)}")

    # Save any remaining records
    if batch_buffer:
        print(f"\nSaving final {len(batch_buffer)} records...")
        write_batch(ws, batch_buffer, next_row)
        total_scraped += len(batch_buffer)
        wb.save(OUTPUT_FILE)

    print(f"\n{'='*60}")
    print(f"DONE! Total medicines scraped: {total_scraped}")
    print(f"File saved: {OUTPUT_FILE}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()