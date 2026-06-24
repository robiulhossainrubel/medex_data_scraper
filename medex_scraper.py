import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import os

OUTPUT_FILE = "medex_medicines.xlsx"
BASE_URL = "https://medex.com.bd/brands"
BATCH_SIZE = 500
DELAY = 1.5  # seconds between requests

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}


def setup_workbook():
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Medicines"

        headers = ["Name", "Type", "Dose", "Generic", "Company", "URL"]
        header_fill = PatternFill("solid", start_color="2E7D32")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_widths = [25, 20, 20, 35, 35, 60]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 20
        next_row = 2

    return wb, ws, next_row


def get_total_pages():
    r = requests.get(BASE_URL, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, "html.parser")
    pages = soup.select("ul.pagination li.page-item a.page-link")
    nums = []
    for a in pages:
        try:
            nums.append(int(a.text.strip()))
        except ValueError:
            pass
    return max(nums) if nums else 1


def parse_medicines(soup):
    medicines = []
    for block in soup.select("a.hoverable-block"):
        try:
            url = block.get("href", "").strip()

            top_div = block.select_one(".data-row-top")
            name = ""
            med_type = ""
            if top_div:
                img = top_div.select_one("img.dosage-icon")
                med_type = img.get("title", "").strip() if img else ""
                name = top_div.get_text(strip=True)
                if img and img.get("alt"):
                    name = name.replace(img.get("alt", ""), "").strip()

            strength_div = block.select_one(".data-row-strength .grey-ligten")
            dose = strength_div.get_text(strip=True) if strength_div else ""

            all_divs = block.select(".data-row > div.col-xs-12")
            generic = ""
            company = ""
            if len(all_divs) >= 3:
                generic_div = all_divs[2]
                if not generic_div.select_one("span"):
                    generic = generic_div.get_text(strip=True)
            if len(all_divs) >= 4:
                company_span = all_divs[3].select_one(".data-row-company")
                company = company_span.get_text(strip=True) if company_span else ""

            medicines.append({
                "name": name,
                "type": med_type,
                "dose": dose,
                "generic": generic,
                "company": company,
                "url": url,
            })
        except Exception as e:
            print(f"  Parse error: {e}")
    return medicines


def write_batch(ws, medicines, start_row):
    row_fills = [
        PatternFill("solid", start_color="F9FBE7"),
        PatternFill("solid", start_color="FFFFFF"),
    ]
    font = Font(name="Arial", size=10)
    for i, med in enumerate(medicines):
        row = start_row + i
        fill = row_fills[i % 2]
        values = [med["name"], med["type"], med["dose"], med["generic"], med["company"], med["url"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))


def main():
    print("Fetching total pages...")
    total_pages = get_total_pages()
    print(f"Total pages: {total_pages}")

    wb, ws, next_row = setup_workbook()
    total_scraped = next_row - 2
    batch_buffer = []

    print(f"Starting scrape. Already have {total_scraped} records.")

    # Calculate which page to start from
    start_page = (total_scraped // 30) + 1  # ~30 items per page
    if total_scraped > 0:
        print(f"Resuming from page {start_page}")

    for page in range(start_page, total_pages + 1):
        url = f"{BASE_URL}?page={page}"
        print(f"Scraping page {page}/{total_pages} | Buffer: {len(batch_buffer)} | Total: {total_scraped}", end="\r")

        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            meds = parse_medicines(soup)
            batch_buffer.extend(meds)

            if len(batch_buffer) >= BATCH_SIZE:
                print(f"\nSaving batch of {BATCH_SIZE} records at row {next_row}...")
                write_batch(ws, batch_buffer[:BATCH_SIZE], next_row)
                next_row += BATCH_SIZE
                total_scraped += BATCH_SIZE
                batch_buffer = batch_buffer[BATCH_SIZE:]
                wb.save(OUTPUT_FILE)
                print(f"Saved! Total records: {total_scraped}")

        except requests.RequestException as e:
            print(f"\nRequest failed on page {page}: {e}. Retrying in 5s...")
            time.sleep(5)
            continue

        time.sleep(DELAY)

    # Save remaining
    if batch_buffer:
        print(f"\nSaving final batch of {len(batch_buffer)} records...")
        write_batch(ws, batch_buffer, next_row)
        total_scraped += len(batch_buffer)
        wb.save(OUTPUT_FILE)

    print(f"\nDone! Total medicines scraped: {total_scraped}")
    print(f"File saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
